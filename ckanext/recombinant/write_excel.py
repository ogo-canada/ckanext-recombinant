import openpyxl

from ckanext.recombinant.tables import get_geno
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import datastore_type
from ckanext.recombinant.helpers import (
    recombinant_choice_fields, recombinant_language_text)

from ckan.plugins.toolkit import _

red_fill = openpyxl.styles.PatternFill(start_color='FFEE1111',
    end_color='FFEE1111', fill_type='solid')

def excel_template(dataset_type, org):
    """
    return an openpyxl.Workbook object containing the sheet and header fields
    for passed dataset_type and org.
    """
    geno = get_geno(dataset_type)

    book = openpyxl.Workbook()
    sheet = book.active
    refs = []
    for chromo in geno['resources']:
        _populate_excel_sheet(sheet, chromo, org, refs)
        sheet = book.create_sheet()

    _populate_reference_sheet(sheet, chromo, refs)
    sheet.title = 'reference'
    return book


def excel_data_dictionary(chromo):
    """
    return an openpyxl.Workbook object containing the field reference
    from chromo, one sheet per language
    """
    book = openpyxl.Workbook()
    sheet = book.active

    style1 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFFFF056'},
        'Font': {
            'bold': True}}
    style2 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFDFE2DB'}}

    from pylons import config
    from ckan.lib.i18n import handle_request, get_lang
    from ckan.common import c, request

    for lang in config['ckan.locales_offered'].split():
        if sheet is None:
            sheet = book.create_sheet()

        sheet.title = lang.upper()
        # switch language (FIXME: this is harder than it should be)
        request.environ['CKAN_LANG'] = lang
        handle_request(request, c)
        choice_fields = dict(
            (f['datastore_id'], f['choices'])
            for f in recombinant_choice_fields(chromo['resource_name']))

        refs = []
        for field in chromo['fields']:
            _append_field_ref_rows(refs, field, style1, style2)

            if field['datastore_id'] in choice_fields:
                _append_field_choices_rows(
                    refs,
                    choice_fields[field['datastore_id']])

        _populate_reference_sheet(sheet, chromo, refs)
        sheet = None

    return book


def _populate_excel_sheet(sheet, chromo, org, refs):
    """
    Format openpyxl sheet for the resource definition chromo and org.

    refs - list of rows to add to reference sheet, modified
        in place from this function

    returns field information for reference sheet
    """
    boolean_validator = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1='"FALSE,TRUE"', allow_blank=True)
    sheet.add_data_validation(boolean_validator)

    sheet.title = chromo['resource_name']

    def fill_cell(row, column, value, styles):
        c = sheet.cell(row=row, column=column)
        c.value = value
        apply_styles(styles, c)

    org_style = chromo['excel_organization_style']
    fill_cell(1, 1, org['name'], org_style)
    fill_cell(1, 2, org['title'], org_style)
    apply_styles(org_style, sheet.row_dimensions[1])

    header_style = chromo['excel_header_style']

    choice_fields = dict(
        (f['datastore_id'], f['choices'])
        for f in recombinant_choice_fields(chromo['resource_name']))

    for n, field in enumerate((f for f in chromo['fields'] if f.get(
            'import_template_include', True)), 1):
        fill_cell(2, n, recombinant_language_text(field['label']), header_style)
        fill_cell(3, n, field['datastore_id'], header_style)
        # jumping through openpyxl hoops:
        col_letter = openpyxl.cell.get_column_letter(n)
        col = sheet.column_dimensions[col_letter]
        col.width = field['excel_column_width']
        # FIXME: format only below header
        col.number_format = datastore_type[field['datastore_type']].xl_format
        validation_range = '{0}4:{0}1004'.format(col_letter)

        _append_field_ref_rows(refs, field, org_style, header_style)

        if field['datastore_type'] == 'boolean':
            boolean_validator.ranges.append(validation_range)
        if field['datastore_id'] in choice_fields:
            ref1 = len(refs) + 1
            _append_field_choices_rows(
                refs,
                choice_fields[field['datastore_id']])
            refN = len(refs)

            if field['datastore_type'] == '_text':
                continue  # can't validate these in excel yet

            choice_range = 'reference!$B${0}:$B${1}'.format(ref1, refN)
            v = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1=choice_range,
                allow_blank=True)
            v.errorTitle = u'Invalid choice'
            v.error = (u'Please enter one of the valid keys shown on '
                'sheet "reference" rows {0}-{1}'.format(ref1, refN))
            sheet.add_data_validation(v)
            v.ranges.append(validation_range)

            # hilight header if bad values pasted below
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                openpyxl.formatting.FormulaRule([(
                    'COUNTIF({0},"<>"&"")' # all non-blank cells
                    '-SUMPRODUCT(COUNTIF({0},{1}))'
                    .format(validation_range, choice_range))],
                    stopIfTrue=True, fill=red_fill))

    apply_styles(header_style, sheet.row_dimensions[2])
    apply_styles(header_style, sheet.row_dimensions[3])
    sheet.row_dimensions[3].hidden = True

    sheet.freeze_panes = sheet['A4']


def _append_field_ref_rows(refs, field, style1, style2):
    refs.append((None, []))
    refs.append((style1, [
        _('Field Name'),
        recombinant_language_text(field['label'])]))
    refs.append((style2, [
        _('ID'),
        field['datastore_id']]))
    if 'description' in field:
        refs.append((style2, [
            _('Description'),
            recombinant_language_text(field['description'])]))
    if 'obligation' in field:
        refs.append((style2, [
            _('Obligation'),
            recombinant_language_text(field['obligation'])]))
    if 'format_type' in field:
        refs.append((style2, [
            _('Format'),
            recombinant_language_text(field['format_type'])]))

def _append_field_choices_rows(refs, choices):
    label = _('Values')
    for key, value in choices:
        refs.append((None, [label, unicode(key), value]))
        label = None

def _populate_reference_sheet(sheet, chromo, refs):
    for style, ref_line in refs:
        sheet.append(ref_line)
        if not style:
            continue

        apply_styles(style, sheet.row_dimensions[sheet.max_row])
        for c in range(len(ref_line)):
            apply_styles(style, sheet.cell(
                row=sheet.max_row, column=c + 1))


def apply_styles(config, target):
    """
    apply styles from config to target

    currently supports PatternFill and Font
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)
